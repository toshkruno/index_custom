import logging, io, base64
import re
import string
from odoo import models, fields, _
from odoo.exceptions import ValidationError
import datetime
from datetime import date
import calendar
from xlwt import easyxf
from odoo.tools.misc import xlwt
from openpyxl.styles import PatternFill, Border, Alignment, Font


_logger = logging.getLogger(__name__)
try:
    import openpyxl
except ImportError:
    msg = _('Install python module "openpyxl" in order to create Excel documents')
    raise ValidationError(msg)
try:
    import csv
except ImportError:
    msg = _('Install python module "csv" in order to generate CSV')
    raise ValidationError(msg)


class salaryAdvanceSummaryReport(models.TransientModel):
    _name = "salary_advance.report"

    months = fields.Selection([('January', 'January'), ('February', 'February'), ('March', 'March'), ('April', 'April'),
                              ('May', 'May'), ('June', 'June'), ('July', 'July'), ('August', 'August'),
                              ('September', 'September'), ('October', 'October'), ('November', 'November'), ('December', 'December')], string='Month', required=True)
    banks = fields.Selection([('Stanbic Bank', 'Stanbic Bank'), ('Co-operative Bank', 'Co-operative Bank')], string='Bank', required=True)
    
    salary_advance_file = fields.Binary('Salary Advance Report')
    file_name = fields.Char('XLS File Name')
    salary_advance_report_printed = fields.Boolean('Salary Advance Report Printed')
    
    def salaryAdvanceReport(self):
        
        if self.banks == 'Co-operative Bank':
            wb = xlwt.Workbook()
            ws = wb.add_sheet("Coop Bank")
            column_heading_style = easyxf('font:height 200;font:bold True;align: horiz left;')

            col = ['Customer Reference No', 'Beneficiary Name', 'Beneficiary Account No.', 'Payment Amount', 'Purpose Of Payment', 'Transaction Currency', 'Charge Type']

            data = []
            ws_count = 0
            for cell, title in zip(list(string.ascii_uppercase), col):
                ws.write(0, ws_count, _(title), column_heading_style)
                ws_count+=1

            for rec in self:
                contracts = rec.env['hr.contract'].search([('state', '=', "open")])

                if contracts:
                    for contract in contracts:
                        salary_advance = self.env['ke.deductions'].search([('employee_id.id', '=', contract.employee_id.id), ('rule_id', '=', rec.env.ref('hr_ke.ke_rule108').id)], limit=1)
                        data.append({
                            'payroll_no': str(contract.employee_id.payroll_no) or '0',
                            'name': contract.employee_id.name or None,
                            'acc_number': contract.employee_id.bank_account_id.acc_number or None,
                            'code': contract.employee_id.bank_account_id.bank_id.bic or None,
                            'amount': salary_advance.amount,
                            'purpose': 'Salary Advance',
                            'currency': 'KES',
                            'charge_type': 'BEN',
                            'bank': contract.employee_id.bank_account_id.bank_id.name or None,
                        })
                else:
                    msg = _('No Contracts to process!')
                    raise ValidationError(msg)

                data = sorted(data, key=lambda i: i['payroll_no'])
                ws_count = 2
                for index, val in enumerate(data):          
                    if val['bank'] and 'Co-operative' in val['bank'] and val['amount'] != 0:
                        ws.write(ws_count, 0, (val['payroll_no'] if val['payroll_no'] != '0' else None or 0),easyxf('font:height 200;align: horiz left;'))
                        ws.write(ws_count, 1, (val['name']),easyxf('font:height 200;align: horiz left;'))
                        ws.write(ws_count, 2, (val['acc_number']),easyxf('font:height 200;align: horiz left;'))
                        ws.write(ws_count, 3, (val['amount']),easyxf('font:height 200;align: horiz left;'))
                        ws.write(ws_count, 4, (val['purpose']),easyxf('font:height 200;align: horiz left;'))
                        ws.write(ws_count, 5, (val['currency']),easyxf('font:height 200;align: horiz left;'))
                        ws.write(ws_count, 6, (val['charge_type']),easyxf('font:height 200;align: horiz left;'))
                        ws_count+=1

                fp = io.BytesIO()
                wb.save(fp)
                excel_file = base64.encodestring(fp.getvalue())
                rec.salary_advance_file = excel_file
                rec.file_name = 'Co-operative Bank(Salary Adavace).xls'
                rec.salary_advance_report_printed = True
                fp.close()
                return {
                        'view_mode': 'form',
                        'res_id': self.id,
                        'res_model': 'salary_advance.report',
                        'view_type': 'form',
                        'type': 'ir.actions.act_window',
                        'context': self.env.context,
                        'target': 'new',
                               }
        
        if self.banks == 'Stanbic Bank':
            wb = xlwt.Workbook()
            ws1 = wb.add_sheet("Stanbic Bank")
            column_heading_style = easyxf('font:height 200;font:bold True;align: horiz left;')

            cols = ['Beneficiary Name', 'Beneficiary Account No.', 'Bank Code', 'Amount', 'Payment Reference']

            data = []
            ws1_count = 0
            for cell, title in zip(list(string.ascii_uppercase), cols):
                ws1.write(0, ws1_count, _(title), column_heading_style)
                ws1_count+=1

            for rec in self:
                contracts = rec.env['hr.contract'].search([('state', '=', "open")])

                if contracts:
                    for contract in contracts:
                        salary_advance = self.env['ke.deductions'].search([('employee_id.id', '=', contract.employee_id.id), ('rule_id', '=', rec.env.ref('hr_ke.ke_rule108').id)], limit=1)
                        data.append({
                            'payroll_no': str(contract.employee_id.payroll_no) or '0',
                            'name': contract.employee_id.name or None,
                            'acc_number': contract.employee_id.bank_account_id.acc_number or None,
                            'code': contract.employee_id.bank_account_id.bank_id.bic or None,
                            'amount': salary_advance.amount,
                            'purpose': 'Salary Advance',
                            'currency': 'KES',
                            'charge_type': 'BEN',
                            'bank': contract.employee_id.bank_account_id.bank_id.name or None,
                        })
                else:
                    msg = _('No Contracts to process!')
                    raise ValidationError(msg)

                data = sorted(data, key=lambda i: i['payroll_no'])
                ws1_count = 2
                for index, val in enumerate(data):
                    if val['bank'] and 'Stanbic' in val['bank'] and val['amount'] != 0:
                        ws1.write(ws1_count, 0, (val['name']),easyxf('font:height 200;align: horiz left;'))
                        ws1.write(ws1_count, 1, (val['acc_number']),easyxf('font:height 200;align: horiz left;'))
                        ws1.write(ws1_count, 2, (val['code']),easyxf('font:height 200;align: horiz left;'))
                        ws1.write(ws1_count, 3, (val['amount']),easyxf('font:height 200;align: horiz left;'))
                        ws1.write(ws1_count, 4, ((rec.months) + ' ' + 'Month' + ' ' + 'Salary'),easyxf('font:height 200;align: horiz left;'))
                        ws1_count+=1

                fp = io.BytesIO()
                wb.save(fp)
                excel_file = base64.encodestring(fp.getvalue())
                rec.salary_advance_file = excel_file
                rec.file_name = 'Stanbic Bank(Salary Adavace).xls'
                rec.salary_advance_report_printed = True
                fp.close()
                return {
                        'view_mode': 'form',
                        'res_id': self.id,
                        'res_model': 'salary_advance.report',
                        'view_type': 'form',
                        'type': 'ir.actions.act_window',
                        'context': self.env.context,
                        'target': 'new',
                               }


class PayrollReports(models.Model):
    _inherit = 'hr.payslip.run'
    
    def nonCashBenefitsReport(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        small_font = openpyxl.styles.Font(size=12.5, bold=True)

        col = ['Payroll No', 'Name', 'Airtime', 'Fuel', 'Medical', 'Mobile Phone', 'Security Services', 'Water', 'Motor Vehicle']
        
        data = []
        for cell, title in zip(list(string.ascii_uppercase), col):
            ws[f'{cell}1'] = title
            ws[f'{cell}1'].font = small_font

        for rec in self:
            if rec.slip_ids:
                filename_summary = 'Non_cash Benefits-' + re.sub(
                    '[^A-Za-z0-9]+', '', rec.name) + '_' + fields.Datetime.context_timestamp(
                    self, fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.xlsx'

                for slip in rec.slip_ids:

                    airtime = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule34').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    fuel = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule35').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    medical = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule31').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    mobile_phone = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule33').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    security_services = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule36').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    water = slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule32').id), ('slip_id', '=', slip.id)], limit=1).total or 0.0
                    car = (slip.line_ids.search([('salary_rule_id', '=', rec.env.ref(
                        'hr_ke.ke_rule38').id), ('slip_id', '=', slip.id)], limit=1).total) if slip.contract_id.car else 0

                    data.append({
                        'payroll_no': str(slip.employee_id.payroll_no) or '0',
                        'name': slip.employee_id.name or None,
                        'airtime': airtime,
                        'fuel': fuel,
                        'medical': medical,
                        'mobile_phone': mobile_phone,
                        'security_services': security_services,
                        'water': water,
                        'car': car,
                    })
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)

        data = sorted(data, key=lambda i: i['payroll_no'])

        for index, val in enumerate(data):
            row = index+2
            ws.cell(row, 1).value = val['payroll_no'] if val['payroll_no'] != '0' else None
            ws.cell(row, 2).value = val['name']
            ws.cell(row, 3).value = val['airtime']
            ws.cell(row, 4).value = val['fuel']
            ws.cell(row, 5).value = val['medical']
            ws.cell(row, 6).value = val['mobile_phone']
            ws.cell(row, 7).value = val['security_services']
            ws.cell(row, 8).value = val['water']
            ws.cell(row, 9).value = val['car']

        for rec in self:
            if rec.slip_ids:
                # Save file as attachment
                xls_path = self.env['hr.ke'].create_xls()
                wb.save(xls_path)
                rec.env['hr.ke'].save_attachment(
                    filename_summary, xls_path, self._name, rec.id)
